"""
Layer 3 — the accuracy eval.

Compares the human answer key (answer_key.xlsx) against the stored profiles,
field by field, and prints the accuracy scoreboard:
  seniority % · industry % · job_family % · dates % · job capture %

Grading rules:
- seniority: exact match per person.
- jobs: truth rows are matched to model jobs per person by title similarity
  (order as tiebreak). Unmatched truth jobs count against job capture.
- industry / job_family: exact id match on matched jobs.
- dates: truth bare year 'YYYY' -> model matching the year counts as correct
  (month unknowable). Truth 'YYYY-MM' -> exact match required. Blank truth
  'to' means current -> model null/blank is correct. Model inventing a date
  where truth is blank counts wrong; model missing a date truth has = wrong.
"""

import re
import sys
import json
from pathlib import Path
from openpyxl import load_workbook

BASE = Path(__file__).parent
KEY = BASE / "answer_key.xlsx"
PROFILES = BASE / "profiles"

# If the key's source_file names differ from the actual profile filenames
# (e.g. the grader used an older template), map old -> real names in a local
# answer_key_aliases.json ({"old name.txt": "real name.txt"}). That file is
# gitignored: it can contain candidate names, this script must not.
ALIAS = {}
_alias_path = BASE / "answer_key_aliases.json"
if _alias_path.exists():
    ALIAS = json.loads(_alias_path.read_text(encoding="utf-8"))


def norm_title(t):
    return set(re.sub(r"[^a-z0-9\u0600-\u06FF ]", " ", (t or "").lower()).split())


def title_sim(a, b):
    A, B = norm_title(a), norm_title(b)
    if not A or not B:
        return 0.0
    return len(A & B) / len(A | B)


def load_key(path):
    wb = load_workbook(path, data_only=True)
    p, j = wb["persons"], wb["jobs"]
    persons, jobs = {}, {}
    for r in range(2, p.max_row + 1):
        sf = p.cell(row=r, column=1).value
        if not sf or str(sf).startswith(("EXAMPLE", "LEGEND")):
            continue
        sf = ALIAS.get(str(sf).strip(), str(sf).strip())
        persons[sf] = str(p.cell(row=r, column=3).value).strip()
    for r in range(2, j.max_row + 1):
        sf = j.cell(row=r, column=1).value
        if not sf or str(sf).startswith("EXAMPLE"):
            continue
        sf = ALIAS.get(str(sf).strip(), str(sf).strip())
        jobs.setdefault(sf, []).append({
            "n": j.cell(row=r, column=2).value,
            "title": str(j.cell(row=r, column=3).value or "").strip(),
            "ind": str(j.cell(row=r, column=5).value or "").strip(),
            "fam": str(j.cell(row=r, column=6).value or "").strip(),
            "from": str(j.cell(row=r, column=7).value or "").strip(),
            "to": str(j.cell(row=r, column=8).value or "").strip(),
        })
    for sf in jobs:
        jobs[sf].sort(key=lambda x: (x["n"] is None, x["n"]))
    return persons, jobs


def date_match(truth, model):
    """Return True/False, or None to skip (both empty)."""
    t = (truth or "").strip()
    m = "" if model in (None, "") else str(model).strip()
    if not t and not m:
        return None                      # both silent (e.g. current job) -> fine, skip
    if not t or not m:
        return False                     # one side invented or missed
    if re.fullmatch(r"\d{4}", t):
        return m[:4] == t                # truth is bare year -> year-level match
    return m == t                        # truth has month -> exact


def match_jobs(truth_list, model_list):
    """Greedy best-similarity matching; returns list of (truth, model|None)."""
    pairs, used = [], set()
    for t in truth_list:
        best, best_s = None, 0.0
        for i, mjob in enumerate(model_list):
            if i in used:
                continue
            s = title_sim(t["title"], mjob.get("title"))
            if s > best_s:
                best, best_s = i, s
        if best is not None and best_s >= 0.25:
            used.add(best)
            pairs.append((t, model_list[best]))
        else:
            pairs.append((t, None))
    extras = [m for i, m in enumerate(model_list) if i not in used]
    return pairs, extras


def main():
    persons, jobs_key = load_key(KEY)
    scores = {"seniority": [0, 0], "industry": [0, 0], "job_family": [0, 0],
              "dates": [0, 0], "capture": [0, 0]}
    sen_rows, miss_rows, extra_total = [], [], 0

    for sf, true_sen in persons.items():
        pf = PROFILES / (sf[:-4] + ".json")
        if not pf.exists():
            print(f"!! profile not found for {sf}")
            continue
        prof = json.loads(pf.read_text(encoding="utf-8"))["profile"]

        model_sen = prof.get("seniority")
        ok = (model_sen == true_sen)
        scores["seniority"][0] += ok
        scores["seniority"][1] += 1
        sen_rows.append((prof.get("full_name") or sf, true_sen, model_sen, "OK" if ok else "X"))

        pairs, extras = match_jobs(jobs_key.get(sf, []), prof.get("experience", []))
        extra_total += len(extras)
        for t, m in pairs:
            scores["capture"][1] += 1
            if m is None:
                miss_rows.append((prof.get("full_name") or sf, t["title"]))
                continue
            scores["capture"][0] += 1
            if t["ind"]:
                scores["industry"][1] += 1
                scores["industry"][0] += (m.get("industry") == t["ind"])
            if t["fam"]:
                scores["job_family"][1] += 1
                scores["job_family"][0] += (m.get("job_family") == t["fam"])
            for tk, mk in (("from", "from"), ("to", "to")):
                r = date_match(t[tk], m.get(mk))
                if r is None:
                    continue
                scores["dates"][1] += 1
                scores["dates"][0] += r

    print("=" * 68)
    print("LAYER 3 — ACCURACY EVAL  (answer key vs stored profiles)")
    print("-" * 68)
    for k in ("seniority", "industry", "job_family", "dates", "capture"):
        hit, tot = scores[k]
        pct = hit / tot * 100 if tot else 0
        print(f"  {k:<12}: {hit:>3}/{tot:<3}  {pct:5.1f}%")
    print("-" * 68)
    print("  seniority per person (truth -> model):")
    for name, t, m, ok in sen_rows:
        print(f"    [{ok:^2}] {str(name)[:30]:<30} {t:<9} -> {m}")
    if miss_rows:
        print("-" * 68)
        print("  truth jobs the model NEVER extracted (capture misses):")
        for name, title in miss_rows:
            print(f"    {str(name)[:28]:<28} | {title[:38]}")
    if extra_total:
        print(f"  (+ {extra_total} model jobs with no truth row — extra/merged entries)")
    print("=" * 68)


if __name__ == "__main__":
    main()
